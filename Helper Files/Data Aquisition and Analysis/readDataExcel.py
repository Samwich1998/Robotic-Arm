#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar  2 13:56:47 2021

@author: samuelsolomon
"""

# General Modules
import os
import re
import numpy as np
# Read/Write to Excel
import openpyxl as xl
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from peakAnalysis import globalParam

# --------------------------------------------------------------------------- #
# --------------------- Extract Test Data from Excel ------------------------ #

class readExcel(globalParam):
    def __init__(self, xWidth = 2000, moveDataFinger = 200, numChannels = 4):
        # Get Variables from Peak Analysis File
        super().__init__(xWidth, moveDataFinger, numChannels)
            
        
    def streamExcelData(self, testDataExcelFile, seeFullPlot = False, testNeuralNetwork = False, 
                        testSheetNum = 0, Controller=None, nn=None, analyzeSheet = None):
        """
        Extracts EMG Data from Excel Document (.xlsx). Data can be iplaced n any,
        worksheet which the user can specify using 'testSheetNum'.
        In the Worksheet:
            Channel 1 Data must be in Column 'A'
            Channel 2 Data must be in Column 'B'
            Channel 3 Data must be in Column 'C'
            Channel 4 Data must be in Column 'D'
        If No Data is present in a cell, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            testDataExcelFile: The Path to the Excel File Containing the Channel Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) of the data.
        --------------------------------------------------------------------------
        """
        # Reset Global Variable in Case it Was Previously Populated
        self.resetGlobalVariables()
    
        # Load Data from Excel File if No Sheet Given
        if analyzeSheet == None:
            print("\nAttempting to Extract Data from the Excel File:", testDataExcelFile)
            WB = xl.load_workbook(testDataExcelFile, data_only=True,read_only=True)
            WB_worksheets = WB.worksheets
            analyzeSheet = WB_worksheets[testSheetNum]
        
        # If Header Exists, Skip Until You Find the Data Data
        for row in analyzeSheet.rows:
            cellA = row[0]
            if type(cellA.value) == type(1.1):
                dataStartRow = cellA.row
                break
        
        dataFinger = 0
        # Loop Through the Excel Worksheet to collect all the data
        for pointNum, [colA, colB, colC, colD] in enumerate(analyzeSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=4, max_row=analyzeSheet.max_row)):
            # Get Cell Values for First 4 Channels: Represents the Voltage for Each Channel;
            V1 = colA.value; V2 = colB.value;
            V3 = colC.value; V4 = colD.value;
            
            # SafeGaurd: If User Edits the Document to Create Empty Rows, Stop Reading in Data
            if V1 == None and V2 == None and V3 == None and V4 == None:
                break
            
            # Add EMG Global Data to Dictionary in Sequential Order.
            self.data["time_ms"].append(pointNum)
            self.data["Channel1"].append(float(colA.value or 0))  # Represent No Value (None) as 0
            self.data["Channel2"].append(float(colB.value or 0))  # Represent No Value (None) as 0
            self.data["Channel3"].append(float(colC.value or 0))  # Represent No Value (None) as 0
            self.data["Channel4"].append(float(colD.value or 0))  # Represent No Value (None) as 0
            
            # When Ready, Send Data Off for Analysis
            while pointNum+1 - dataFinger >= self.xWidth:
                self.live_plotter(dataFinger, seeFullPlot, testNeuralNetwork, Controller=Controller, nn=nn)
                dataFinger += self.moveDataFinger
            
        # Finished Data Collection: Report Back to User
        print("\tDone Data Collecting from File: ", analyzeSheet.title)
        
    
    def getDataFromTrainingFile(self, excelSheet, Training_Data, Training_Labels, movementOptions):
        # Get Hand Movement
        currentLabel = excelSheet.title.split(" - ")[1]
        featureLabel = movementOptions == currentLabel.lower()
        featureLabel = featureLabel.astype(int)
        
        # If Header Exists, Skip Until You Find the Data Data
        for row in excelSheet.rows:
            cellI = row[8]; cellJ = row[9]
            cellK = row[10]; cellL = row[11]
            if type(cellI.value) == type(1.1) or type(cellJ.value) == type(1.1) or type(cellK.value) == type(1.1) or type(cellL.value) == type(1.1):
                dataStartRow = row[0].row
                break
        
        # Create Data Structure to Hold Results Until Full Group is Present
        dataHold = np.empty((0, self.numChannels), float)
                
        # Loop Through the Excel Worksheet to collect all the data
        for (colI, colJ, colK, colL) in excelSheet.iter_rows(min_col=9, min_row=dataStartRow, max_col=9+self.numChannels-1, max_row=excelSheet.max_row):
            # Get Cell Values for First 4 Channels: Represents the Voltage for Each Channel;
            feature1 = colI.value; feature2 = colJ.value;
            feature3 = colK.value; feature4 = colL.value;
            
            # If All rows are empty, it is a new group
            if feature1 == None and feature2 == None and feature3 == None and feature4 == None:
                # Base Case: dataHold didnt change, we are done collecting
                if len(dataHold) == 0:
                    break
                # Collect Point as Data
                featureAv = np.nanmean(np.where(dataHold>0, dataHold, np.nan), axis=0)
                Training_Data = np.vstack((Training_Data, featureAv))
                Training_Labels = np.vstack((Training_Labels, featureLabel))
                # Reset Group Holder
                dataHold = np.empty((0, self.numChannels), float)
                continue
            
            # Collect Features in Cirect Group
            featureList = [float(feature1 or 0), float(feature2 or 0), float(feature3 or 0), float(feature4 or 0)]
            dataHold = np.vstack((dataHold, featureList))
        
        print("\tCollected Training Data for:", excelSheet.title)
        return Training_Data, Training_Labels
    
    def getTrainingData(self, trainDataExcelFolder, movementOptions, mode):
        if mode == 'Train':
            NN_Data = np.empty((0, self.numChannels), float)
            NN_Labels = np.empty((0, len(movementOptions)), float)
            
        for excelFile in list(os.listdir(trainDataExcelFolder)):
            if excelFile.endswith(".xlsx") and not excelFile.startswith("~"):
                # Get Full Path to the Excel File
                trainingExcelFile = trainDataExcelFolder + excelFile
                print("\nLoading Excel File", trainingExcelFile)
                # Load the Excel File
                if mode == 'Train':
                    WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=True)
                else:
                    WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=False)
                WB_worksheets = WB.worksheets
                # Loop Over Each Sheet in the File
                saveExcelData = saveExcel(self.numChannels)
                for excelSheet in WB_worksheets:
                    self.resetGlobalVariables()
                    
                    # Get the Training Data/Label from the Sheet
                    if mode == 'Train':
                        NN_Data, NN_Labels = self.getDataFromTrainingFile(excelSheet, NN_Data, NN_Labels, movementOptions)
                    elif mode == 'reAnalyze':
                        print("\tReanalyzing Excel Sheet:", excelSheet.title)
                        # Delete the Previous Analysis from Excel
                        excelSheet.delete_cols(5,8) # Delete Columns E-L (5 -> 5+8-1)
                        # ReAnalyze Data (First Four Columns)
                        self.streamExcelData(trainingExcelFile, analyzeSheet=excelSheet)
                        # Overwrite Excel Sheet with new Analysis
                        handMovement = excelSheet.title.split(" - ")[1]
                        saveExcelData.saveTestData(self.data, self.xTopGrouping, self.featureSetGrouping, trainDataExcelFolder, excelFile, sheetName=excelSheet, handMovement=handMovement, overWrite = True)
                        
        if mode == 'Train':
            NaN_Placements = np.isnan(NN_Data)
            NN_Data[NaN_Placements] = 0
            return NN_Data, NN_Labels
        elif mode == 'reAnalyze':
            return None



class saveExcel:
    def __init__(self, numChannels = 4):
        # Input Parameters
        self.numChannels = numChannels        # Number of EMG Signals
        
        # Specify OpenPyxl Asthetics
        self.openpyxlColors = {
            0: 45,
            1: 46,
            2: 47,
            3: 49,
            4: 50,
            5: 51,
            6: 55,
            }

    def saveData(self, data, xTopGrouping, featureSetGrouping, saveDataFolder, saveExcelName,
                     sheetName = "Trial 1 - No Gesture", handMovement="No Movement", overWrite=False):        
        # Create Output File Directory to Save Data: If None
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excel_file = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excel_file):
            print("\nSaving the Data as New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        # Overwrite Previous Excel Sheet; Replacing Sheetname that was Edited
        elif overWrite:
            print("\tOverWriting Excel File:", excel_file)
            WB = xl.load_workbook(excel_file, read_only=False)
            WB_worksheet = sheetName
        # Loading in Previous Excel File, Creating New Sheet, Editing Trial Number in SheetName
        else:
            print("Excel File Already Exists. Loading File")
            WB = xl.load_workbook(excel_file, read_only=False)
            currentSheets = WB.sheetnames
            # Get All Sheets with the Current Movement
            currentMovementSheets = []
            for sheet in currentSheets:
                movement = sheet.split(" - ")
                if handMovement == movement[1]:
                    currentMovementSheets.append(sheet)
            # Get the Last Trial for this Hand Movement
            sheetName = max(currentMovementSheets, key=lambda x: int(re.findall(r'\d+', x.split(" - ")[0])[0]), default= "Trial 0 - " + handMovement)
            # Edit SheetName
            sheetInfo = sheetName.split(" - ")
            currentTrial = re.findall(r'\d+', sheetInfo[0])[0]
            newTrial = sheetInfo[0].split(currentTrial)[0] + str(int(currentTrial)+1)
            sheetName = newTrial + " - " + sheetInfo[1]
            WB_worksheet = WB.create_sheet(sheetName)
            print("Saving Sheet as", sheetName)
        
        xPeakHeader = ['Channel 1 X Peaks', 'Channel 2 X Peaks', 'Channel 3 X Peaks', 'Channel 4 X Peaks']
        featureHeader = ['Channel 1 Features', 'Channel 2 Features', 'Channel 3 Features', 'Channel 4 Features']
        if not overWrite:
            # Label First Row
            header = ['Channel 1', 'Channel 2', 'Channel 3', 'Channel 4']
            header.extend(xPeakHeader)
            header.extend(featureHeader)
            WB_worksheet.append(header)
            
            # Save Data to Worksheet
            for dataNum in range(len(data['time_ms'])):
                row = []
                for channel in range(self.numChannels):
                    row.append(data['Channel'+str(1+channel)][dataNum])
                WB_worksheet.append(row)
        else:
            for cells in WB_worksheet.iter_rows(min_col=5, min_row=1, max_col=12, max_row=1):
                colE, colF, colG, colH, colI, colJ, colK, colL = cells[0], cells[1], cells[2], cells[3], cells[4], cells[5], cells[6], cells[7]
                [colE.value, colF.value, colG.value, colH.value] = xPeakHeader
                [colI.value, colJ.value, colK.value, colL.value] = featureHeader
                
      #  for col in WB_worksheet.iter_cols(min_row=1, min_col = 1, max_col=3, max_row=1):
       #    print(col, len(col))
        #    col.sheetColsFeatures['A'].width = len(col)
            
        
        # Add X Peaks and Features
        sheetColsXPeaks = ['E','F','G','H']
        sheetColsFeatures = ['I','J','K','L']
        for channel in range(self.numChannels):
            startIndex = 1
            for peakNum in xTopGrouping[channel]:
                rowIndex = startIndex
                peakColor = (peakNum-1)%(len(self.openpyxlColors))
                cellColor = Color(indexed = self.openpyxlColors[peakColor])
                for peakVal in xTopGrouping[channel][peakNum]:
                    WB_worksheet[sheetColsXPeaks[channel]][rowIndex].value = peakVal
                    WB_worksheet[sheetColsXPeaks[channel]][rowIndex].fill = PatternFill(fgColor=cellColor, fill_type = 'solid')
                    WB_worksheet[sheetColsXPeaks[channel]][rowIndex].alignment = Alignment(horizontal='center')
                    rowIndex += 1
                # Set the Same Row Index for All
                startIndex += 1 + len(max(xTopGrouping.values(), key = lambda x: len(x[peakNum]))[peakNum])

        for channel in range(self.numChannels):
            startIndex = 1
            for peakNum in featureSetGrouping[channel]:
                rowIndex = startIndex
                peakColor = (peakNum-1)%(len(self.openpyxlColors))
                cellColor = Color(indexed = self.openpyxlColors[peakColor])
                for featureVal in featureSetGrouping[channel][peakNum]:
                    WB_worksheet[sheetColsFeatures[channel]][rowIndex].value = featureVal
                    WB_worksheet[sheetColsFeatures[channel]][rowIndex].fill = PatternFill(fgColor=cellColor, fill_type = 'solid')
                    WB_worksheet[sheetColsFeatures[channel]][rowIndex].alignment = Alignment(horizontal='center')
                    rowIndex += 1
                startIndex += 1 + len(max(xTopGrouping.values(), key = lambda x: len(x[peakNum]))[peakNum])

        # Save as New Excel File
        WB.save(excel_file)
        WB.close()
        print("DONE")